# Module --> pip install openpyxl
import os 
import openpyxl
os.chdir('.')
workbook = openpyxl.load_workbook('example.xlsx')

print(type(workbook)) # <class 'openpyxl.workbook.workbook.Workbook'>

print(workbook.get_sheet_names())

sheet = workbook.get_sheet_by_name('Hoja1')
print(type(sheet)) #<class 'openpyxl.worksheet.worksheet.Worksheet'>

cell = sheet['A1']
print(cell) # <Cell 'Hoja1'.A1>

print(cell.value) # 35

print(sheet.cell(row=1, column=1)) # <Cell 'Hoja1'.A1>

wb = openpyxl.Workbook() # Creates a workbook object
sheet = wb['sheet']
sheet['A1'] = 42
#wb.save('example.xslx') --> saves excel document

sheet1 = wb.create_sheet(index=1, title='One sheet') #--> Creates a new sheet

sheet1.title('Examplechangesheetname')
